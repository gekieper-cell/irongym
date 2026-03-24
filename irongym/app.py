"""
IRON GYM — Sistema de Gestión de Gimnasio de Boxeo
Flask + SQLite | Windows | Red Local
"""
from flask import Flask, render_template, request, jsonify, send_file, redirect, url_for, session
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime, date, timedelta
from functools import wraps
import os, io, csv, json, base64, hashlib

# ── Entorno: TEST o PRODUCCION ───────────────────────────────────
# Se define con la variable de entorno IRONGYM_ENV
# test        → usa instance_test/irongym_test.db  puerto 5001
# produccion  → usa instance/irongym.db             puerto 5000 (default)
ENTORNO = os.environ.get('IRONGYM_ENV', 'produccion').lower()
IS_TEST = (ENTORNO == 'test')

# ── Pure-Python QR matrix (no Pillow needed) ──────────────────────
def _qr_matrix(data: str):
    """Return a qrcode.QRCode matrix using only the qrcode library (no image deps)."""
    import qrcode as _qr
    qr = _qr.QRCode(version=None, error_correction=_qr.constants.ERROR_CORRECT_M,
                    box_size=1, border=2)
    qr.add_data(data)
    qr.make(fit=True)
    return qr.get_matrix()          # list of list of bool

def _matrix_to_svg(matrix) -> str:
    size = len(matrix)
    cell = 8
    px = size * cell
    rects = []
    for r, row in enumerate(matrix):
        for c, val in enumerate(row):
            if val:
                x, y = c * cell, r * cell
                rects.append(f'<rect x="{x}" y="{y}" width="{cell}" height="{cell}"/>')
    return (f'<svg xmlns="http://www.w3.org/2000/svg" width="{px}" height="{px}" '
            f'viewBox="0 0 {px} {px}" shape-rendering="crispEdges">'
            f'<rect width="100%" height="100%" fill="white"/>'
            f'<g fill="black">{"".join(rects)}</g></svg>')

def gen_qr_svg(data_dict: dict) -> str:
    """Return inline SVG string for a QR code."""
    matrix = _qr_matrix(json.dumps(data_dict))
    return _matrix_to_svg(matrix)

def gen_qr_b64_svg(data_dict: dict) -> str:
    """Return base64-encoded SVG (for use in <img src='data:...'/>)."""
    svg = gen_qr_svg(data_dict)
    return base64.b64encode(svg.encode()).decode()

# ── App config ──────────────────────────────────────────────────
APP_VERSION = "2.1"
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

if IS_TEST:
    DB_FOLDER = os.path.join(BASE_DIR, 'instance_test')
    DB_FILE   = os.path.join(DB_FOLDER, 'irongym_test.db')
    APP_PORT  = 5001
else:
    DB_FOLDER = os.path.join(BASE_DIR, 'instance')
    DB_FILE   = os.path.join(DB_FOLDER, 'irongym.db')
    APP_PORT  = 5000

os.makedirs(DB_FOLDER, exist_ok=True)
app = Flask(__name__, instance_path=DB_FOLDER)
app.secret_key = f'irongym-secret-2024-{ENTORNO}'
app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{DB_FILE}'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
# Producción: pool optimizado para múltiples usuarios simultáneos
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    'connect_args': {
        'check_same_thread': False,
        'timeout': 30,
    },
    'pool_size': 10,
    'pool_timeout': 20,
    'pool_recycle': 300,
}

# Activar WAL mode en SQLite al crear conexión (permite lecturas simultáneas)
from sqlalchemy import event
from sqlalchemy.engine import Engine
import sqlite3

@event.listens_for(Engine, 'connect')
def set_sqlite_pragma(dbapi_conn, connection_record):
    if isinstance(dbapi_conn, sqlite3.Connection):
        cursor = dbapi_conn.cursor()
        cursor.execute('PRAGMA journal_mode=WAL')   # múltiples lectores simultáneos
        cursor.execute('PRAGMA synchronous=NORMAL')  # buen balance velocidad/seguridad
        cursor.execute('PRAGMA cache_size=-64000')   # 64MB cache en RAM
        cursor.execute('PRAGMA foreign_keys=ON')
        cursor.execute('PRAGMA busy_timeout=30000')  # esperar 30s antes de "db locked"
        cursor.close()

db = SQLAlchemy(app)

# ══════════════════════════════════════════════════════════════════
# MODELS
# ══════════════════════════════════════════════════════════════════
class Config(db.Model):
    id      = db.Column(db.Integer, primary_key=True)
    clave   = db.Column(db.String(80), unique=True, nullable=False)
    valor   = db.Column(db.Text, default='')

class Plan(db.Model):
    id      = db.Column(db.Integer, primary_key=True)
    nombre  = db.Column(db.String(120), nullable=False)
    precio  = db.Column(db.Float, default=0)
    activo  = db.Column(db.Boolean, default=True)

class Alumno(db.Model):
    id          = db.Column(db.Integer, primary_key=True)
    nombre      = db.Column(db.String(200), nullable=False)
    dni         = db.Column(db.String(20), unique=True, nullable=False)
    tel         = db.Column(db.String(40))
    email       = db.Column(db.String(120))
    fnac        = db.Column(db.Date)
    alta        = db.Column(db.Date, default=date.today)
    vto_cuota   = db.Column(db.Date)
    plan_id     = db.Column(db.Integer, db.ForeignKey('plan.id'))
    plan        = db.relationship('Plan', backref='alumnos')
    estado      = db.Column(db.String(20), default='activo')   # activo/inactivo/moroso
    monto       = db.Column(db.Float, default=0)
    obs         = db.Column(db.Text)
    created_at  = db.Column(db.DateTime, default=datetime.now)

    @property
    def es_moroso(self):
        if self.estado == 'inactivo': return False
        if not self.vto_cuota: return False
        return self.vto_cuota < date.today()

    @property
    def dias_vencido(self):
        if not self.es_moroso: return 0
        return (date.today() - self.vto_cuota).days

    def to_dict(self):
        return {
            'id': self.id, 'nombre': self.nombre, 'dni': self.dni,
            'tel': self.tel or '', 'email': self.email or '',
            'fnac': self.fnac.isoformat() if self.fnac else '',
            'alta': self.alta.isoformat() if self.alta else '',
            'vto_cuota': self.vto_cuota.isoformat() if self.vto_cuota else '',
            'plan_id': self.plan_id, 'plan_nombre': self.plan.nombre if self.plan else '',
            'estado': self.estado, 'monto': self.monto, 'obs': self.obs or '',
            'es_moroso': self.es_moroso, 'dias_vencido': self.dias_vencido,
        }

class Asistencia(db.Model):
    id          = db.Column(db.Integer, primary_key=True)
    alumno_id   = db.Column(db.Integer, db.ForeignKey('alumno.id'), nullable=False)
    alumno      = db.relationship('Alumno', backref='asistencias')
    fecha       = db.Column(db.Date, default=date.today)
    hora        = db.Column(db.String(8))
    metodo      = db.Column(db.String(10), default='dni')  # qr/dni

    def to_dict(self):
        return {
            'id': self.id, 'alumno_id': self.alumno_id,
            'nombre': self.alumno.nombre if self.alumno else '',
            'dni': self.alumno.dni if self.alumno else '',
            'fecha': self.fecha.isoformat(), 'hora': self.hora, 'metodo': self.metodo,
        }

class Clase(db.Model):
    id          = db.Column(db.Integer, primary_key=True)
    nombre      = db.Column(db.String(120), nullable=False)
    instructor  = db.Column(db.String(120))
    dia         = db.Column(db.String(20))
    hora        = db.Column(db.String(8))
    duracion    = db.Column(db.Integer, default=60)
    cupo        = db.Column(db.Integer, default=20)
    desc        = db.Column(db.Text)
    activa      = db.Column(db.Boolean, default=True)

    def to_dict(self):
        return { 'id': self.id, 'nombre': self.nombre, 'instructor': self.instructor,
                 'dia': self.dia, 'hora': self.hora, 'duracion': self.duracion,
                 'cupo': self.cupo, 'desc': self.desc or '', 'activa': self.activa }

class Producto(db.Model):
    id          = db.Column(db.Integer, primary_key=True)
    nombre      = db.Column(db.String(120), nullable=False)
    categoria   = db.Column(db.String(40))
    emoji       = db.Column(db.String(8), default='📦')
    precio      = db.Column(db.Float, default=0)
    stock       = db.Column(db.Integer, default=0)
    stock_min   = db.Column(db.Integer, default=5)
    desc        = db.Column(db.Text)
    activo      = db.Column(db.Boolean, default=True)

    def to_dict(self):
        return { 'id': self.id, 'nombre': self.nombre, 'categoria': self.categoria,
                 'emoji': self.emoji, 'precio': self.precio, 'stock': self.stock,
                 'stock_min': self.stock_min, 'desc': self.desc or '', 'activo': self.activo }

class Venta(db.Model):
    id          = db.Column(db.Integer, primary_key=True)
    fecha       = db.Column(db.Date, default=date.today)
    hora        = db.Column(db.String(8))
    cliente     = db.Column(db.String(200), default='—')
    total       = db.Column(db.Float, default=0)
    medio_pago  = db.Column(db.String(40))
    items_json  = db.Column(db.Text)
    created_by  = db.Column(db.String(20))

    @property
    def items(self):
        try: return json.loads(self.items_json or '[]')
        except: return []

    def to_dict(self):
        return { 'id': self.id, 'fecha': self.fecha.isoformat(), 'hora': self.hora,
                 'cliente': self.cliente, 'total': self.total, 'medio_pago': self.medio_pago,
                 'items': self.items, 'created_by': self.created_by }

class Gasto(db.Model):
    id          = db.Column(db.Integer, primary_key=True)
    fecha       = db.Column(db.Date, default=date.today)
    desc        = db.Column(db.String(300), nullable=False)
    categoria   = db.Column(db.String(60))
    monto       = db.Column(db.Float, default=0)
    comprobante = db.Column(db.String(200))
    created_by  = db.Column(db.String(20))

    def to_dict(self):
        return { 'id': self.id, 'fecha': self.fecha.isoformat(), 'desc': self.desc,
                 'categoria': self.categoria, 'monto': self.monto,
                 'comprobante': self.comprobante or '', 'created_by': self.created_by }

# ══════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════
def get_cfg(clave, default=''):
    c = Config.query.filter_by(clave=clave).first()
    return c.valor if c else default

def set_cfg(clave, valor):
    c = Config.query.filter_by(clave=clave).first()
    if c: c.valor = valor
    else: db.session.add(Config(clave=clave, valor=valor))
    db.session.commit()

def hash_pass(p): return hashlib.sha256(p.encode()).hexdigest()

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user' not in session:
            return jsonify({'error': 'No autenticado'}), 401
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if session.get('role') != 'admin':
            return jsonify({'error': 'Acceso denegado'}), 403
        return f(*args, **kwargs)
    return decorated


def run_migrations():
    """
    Safe schema migrations — never destroys existing data.
    Add new ALTER TABLE statements here for each version update.
    SQLite ignores 'duplicate column' errors so it's safe to re-run.
    """
    import sqlite3
    db_path = DB_FILE
    if not os.path.exists(db_path):
        return  # fresh install, create_all() already handled it

    conn = sqlite3.connect(db_path)
    cur  = conn.cursor()

    # Helper: add column only if it doesn't exist yet
    def add_column(table, column, col_type, default=None):
        try:
            if default is not None:
                cur.execute(f"ALTER TABLE {table} ADD COLUMN {column} {col_type} DEFAULT {default}")
            else:
                cur.execute(f"ALTER TABLE {table} ADD COLUMN {column} {col_type}")
            conn.commit()
            print(f"  [migración] {table}.{column} agregado")
        except sqlite3.OperationalError:
            pass  # column already exists — no problem

    # ── Migrations list ─────────────────────────────────────────
    # v1.1 — (initial schema via create_all)
    # v1.2 — example: add 'foto' to alumno
    # add_column('alumno', 'foto', 'TEXT')
    #
    # v1.3 — example: add 'descuento' to venta
    # add_column('venta', 'descuento', 'REAL', 0)
    #
    # Para agregar una columna nueva en el futuro:
    # 1. Agregá el modelo en la clase Python arriba
    # 2. Agregá la línea add_column() acá
    # 3. Distribuí el update_X.X.zip
    # ────────────────────────────────────────────────────────────

    conn.close()

    # Save current version to DB config
    v = Config.query.filter_by(clave='app_version').first()
    if v:
        v.valor = APP_VERSION
    else:
        db.session.add(Config(clave='app_version', valor=APP_VERSION))
    db.session.commit()
    print(f"  [sistema] Versión {APP_VERSION} — migraciones OK")

def seed_defaults():
    """Insert default data on first run"""
    if not Config.query.filter_by(clave='pass_admin').first():
        defaults = [
            Config(clave='pass_admin',  valor=hash_pass('admin123')),
            Config(clave='pass_op',     valor=hash_pass('op123')),
            Config(clave='master_key',  valor=hash_pass('IRONGYM2024')),
            Config(clave='gym_nombre',  valor='Iron Gym'),
            Config(clave='gym_dir',     valor=''),
            Config(clave='gym_tel',     valor=''),
            Config(clave='gym_email',   valor=''),
            Config(clave='smtp_host',   valor='smtp.gmail.com'),
            Config(clave='smtp_port',   valor='587'),
            Config(clave='smtp_user',   valor=''),
            Config(clave='smtp_pass',   valor=''),
            Config(clave='fondo_tipo',  valor='color'),   # color | imagen
            Config(clave='fondo_valor', valor='#0a0a0a'), # color hex o URL base64
        ]
        db.session.add_all(defaults)

    if not Plan.query.first():
        planes = [
            Plan(nombre='3 días/semana', precio=8000),
            Plan(nombre='5 días/semana', precio=12000),
            Plan(nombre='Libre',         precio=15000),
            Plan(nombre='Personalizado', precio=20000),
        ]
        db.session.add_all(planes)

    if not Producto.query.first():
        productos = [
            Producto(nombre='Guantes 12oz',     categoria='equipamiento', emoji='🥊', precio=8500, stock=15, stock_min=3),
            Producto(nombre='Guantes 16oz',     categoria='equipamiento', emoji='🥊', precio=9500, stock=10, stock_min=3),
            Producto(nombre='Protector Bucal',  categoria='equipamiento', emoji='🦷', precio=1200, stock=20, stock_min=5),
            Producto(nombre='Cinta de Boxeo',   categoria='equipamiento', emoji='🩹', precio=600,  stock=30, stock_min=10),
            Producto(nombre='Agua 500ml',       categoria='bebidas',      emoji='💧', precio=350,  stock=50, stock_min=10),
            Producto(nombre='Gatorade',         categoria='bebidas',      emoji='🟡', precio=650,  stock=24, stock_min=6),
            Producto(nombre='Powerade',         categoria='bebidas',      emoji='🔵', precio=600,  stock=24, stock_min=6),
            Producto(nombre='Isotónico Premium',categoria='rehidratantes',emoji='⚡', precio=800,  stock=20, stock_min=5),
            Producto(nombre='Proteína Whey 30g',categoria='rehidratantes',emoji='💪', precio=1800, stock=15, stock_min=3),
            Producto(nombre='Electrolitos',     categoria='rehidratantes',emoji='🧪', precio=500,  stock=30, stock_min=8),
        ]
        db.session.add_all(productos)

    if not Clase.query.first():
        clases = [
            Clase(nombre='Boxeo Amateur',    instructor='Coach Rivera', dia='lunes',    hora='08:00', duracion=60, cupo=20),
            Clase(nombre='Sparring',         instructor='Coach Rivera', dia='miercoles',hora='10:00', duracion=90, cupo=10),
            Clase(nombre='Técnica Avanzada', instructor='Coach Méndez', dia='viernes',  hora='18:00', duracion=60, cupo=15),
            Clase(nombre='Cardio Box',       instructor='Coach Méndez', dia='martes',   hora='07:00', duracion=45, cupo=25),
        ]
        db.session.add_all(clases)

    db.session.commit()

# ══════════════════════════════════════════════════════════════════
# AUTH ROUTES
# ══════════════════════════════════════════════════════════════════
@app.route('/')
def index():
    if 'user' in session:
        return render_template('index.html')
    return render_template('login.html')

@app.route('/api/version')
def api_version():
    return jsonify({
        'version':     APP_VERSION,
        'entorno':     ENTORNO,
        'is_test':     IS_TEST,
        'gym_nombre':  get_cfg('gym_nombre', 'Iron Gym'),
        'fondo_tipo':  get_cfg('fondo_tipo','color'),
        'fondo_valor': get_cfg('fondo_valor','#0a0a0a'),
    })

@app.route('/api/login', methods=['POST'])
def api_login():
    data = request.json
    u = (data.get('usuario') or '').lower().strip()
    p = hash_pass(data.get('password') or '')
    if u in ('admin', 'administrador') and p == get_cfg('pass_admin'):
        session['user'] = 'admin'; session['role'] = 'admin'
        return jsonify({'ok': True, 'role': 'admin', 'display': 'Administrador'})
    if u in ('operador', 'op') and p == get_cfg('pass_op'):
        session['user'] = 'operador'; session['role'] = 'operador'
        return jsonify({'ok': True, 'role': 'operador', 'display': 'Operador'})
    return jsonify({'ok': False, 'msg': 'Usuario o contraseña incorrectos'}), 401

@app.route('/api/logout', methods=['POST'])
def api_logout():
    session.clear(); return jsonify({'ok': True})

@app.route('/api/recover/verify', methods=['POST'])
def api_recover_verify():
    key = hash_pass(request.json.get('master_key',''))
    if key == get_cfg('master_key'):
        session['recover_ok'] = True
        return jsonify({'ok': True})
    return jsonify({'ok': False, 'msg': 'Clave maestra incorrecta'}), 401

@app.route('/api/recover/save', methods=['POST'])
def api_recover_save():
    if not session.get('recover_ok'):
        return jsonify({'ok': False}), 403
    p1 = request.json.get('pass1','')
    p2 = request.json.get('pass2','')
    if not p1 or p1 != p2:
        return jsonify({'ok': False, 'msg': 'Contraseñas no coinciden'}), 400
    set_cfg('pass_admin', hash_pass(p1))
    session.pop('recover_ok', None)
    return jsonify({'ok': True})

# ══════════════════════════════════════════════════════════════════
# DASHBOARD
# ══════════════════════════════════════════════════════════════════
@app.route('/api/dashboard')
@login_required
def api_dashboard():
    hoy = date.today()
    mes_ini = hoy.replace(day=1)
    morosos = [a for a in Alumno.query.filter(Alumno.estado != 'inactivo').all() if a.es_moroso]
    ventas_mes = Venta.query.filter(Venta.fecha >= mes_ini).all()
    gastos_mes = Gasto.query.filter(Gasto.fecha >= mes_ini).all()
    asist_hoy  = Asistencia.query.filter(Asistencia.fecha == hoy).count()
    dias = [hoy - timedelta(days=i) for i in range(6,-1,-1)]
    asist_semana = [Asistencia.query.filter(Asistencia.fecha == d).count() for d in dias]
    dias_label   = [d.strftime('%a') for d in dias]
    meses_label, ing_list, gas_list = [], [], []
    for i in range(5,-1,-1):
        d = date(hoy.year, hoy.month, 1) - timedelta(days=i*28)
        d = d.replace(day=1)
        key = d.strftime('%Y-%m')
        meses_label.append(d.strftime('%b'))
        ing_list.append(sum(v.total for v in Venta.query.filter(db.func.strftime('%Y-%m', Venta.fecha)==key).all()))
        gas_list.append(sum(g.monto for g in Gasto.query.filter(db.func.strftime('%Y-%m', Gasto.fecha)==key).all()))
    cats = {}
    for v in ventas_mes:
        for it in v.items:
            c = it.get('cat','otros')
            cats[c] = cats.get(c,0) + it.get('subtotal',0)
    clases_hoy_n = Clase.query.filter(Clase.dia == hoy.strftime('%A').lower()).count()
    dias_es = {'monday':'lunes','tuesday':'martes','wednesday':'miercoles','thursday':'jueves',
               'friday':'viernes','saturday':'sabado','sunday':'domingo'}
    clases_hoy_n = Clase.query.filter(Clase.dia == dias_es.get(hoy.strftime('%A').lower(),'')).count()
    return jsonify({
        'alumnos_activos': Alumno.query.filter_by(estado='activo').count(),
        'morosos': len(morosos),
        'asist_hoy': asist_hoy,
        'clases_hoy': clases_hoy_n,
        'ingresos_mes': sum(v.total for v in ventas_mes),
        'gastos_mes':   sum(g.monto for g in gastos_mes),
        'chart_ig': {'labels': meses_label, 'ingresos': ing_list, 'gastos': gas_list},
        'chart_asist': {'labels': dias_label, 'data': asist_semana},
        'chart_cats': {'labels': list(cats.keys()), 'data': list(cats.values())},
        'chart_estados': {
            'labels': ['Activos','Inactivos','Morosos'],
            'data': [Alumno.query.filter_by(estado='activo').count(),
                     Alumno.query.filter_by(estado='inactivo').count(), len(morosos)]
        },
        'top_morosos': [a.to_dict() for a in morosos[:5]],
    })

# ══════════════════════════════════════════════════════════════════
# ALUMNOS
# ══════════════════════════════════════════════════════════════════
@app.route('/api/alumnos', methods=['GET'])
@login_required
def api_alumnos_list():
    q     = request.args.get('q','').lower()
    est   = request.args.get('estado','')
    query = Alumno.query
    if q:
        query = query.filter(db.or_(Alumno.nombre.ilike(f'%{q}%'), Alumno.dni.ilike(f'%{q}%')))
    if est == 'moroso':
        alumnos = [a for a in query.all() if a.es_moroso]
    elif est:
        alumnos = query.filter_by(estado=est).all()
    else:
        alumnos = query.order_by(Alumno.nombre).all()
    return jsonify([a.to_dict() for a in alumnos])

@app.route('/api/alumnos', methods=['POST'])
@login_required
def api_alumno_create():
    d = request.json
    if Alumno.query.filter_by(dni=d['dni']).first():
        return jsonify({'error': 'Ya existe un alumno con ese DNI'}), 400
    a = Alumno(
        nombre=d['nombre'], dni=d['dni'], tel=d.get('tel'), email=d.get('email'),
        fnac=date.fromisoformat(d['fnac']) if d.get('fnac') else None,
        alta=date.fromisoformat(d['alta']) if d.get('alta') else date.today(),
        vto_cuota=date.fromisoformat(d['vto_cuota']) if d.get('vto_cuota') else None,
        plan_id=d.get('plan_id'), estado=d.get('estado','activo'),
        monto=float(d.get('monto',0)), obs=d.get('obs',''),
    )
    db.session.add(a); db.session.commit()
    return jsonify(a.to_dict()), 201

@app.route('/api/alumnos/<int:id>', methods=['PUT'])
@login_required
def api_alumno_update(id):
    a = Alumno.query.get_or_404(id); d = request.json
    dup = Alumno.query.filter(Alumno.dni==d['dni'], Alumno.id!=id).first()
    if dup: return jsonify({'error': 'DNI duplicado'}), 400
    a.nombre=d['nombre']; a.dni=d['dni']; a.tel=d.get('tel'); a.email=d.get('email')
    a.fnac=date.fromisoformat(d['fnac']) if d.get('fnac') else None
    a.alta=date.fromisoformat(d['alta']) if d.get('alta') else a.alta
    a.vto_cuota=date.fromisoformat(d['vto_cuota']) if d.get('vto_cuota') else None
    a.plan_id=d.get('plan_id'); a.estado=d.get('estado','activo')
    a.monto=float(d.get('monto',0)); a.obs=d.get('obs','')
    db.session.commit(); return jsonify(a.to_dict())

@app.route('/api/alumnos/<int:id>', methods=['DELETE'])
@login_required
@admin_required
def api_alumno_delete(id):
    a = Alumno.query.get_or_404(id); db.session.delete(a); db.session.commit()
    return jsonify({'ok': True})

@app.route('/api/alumnos/<int:id>/qr')
@login_required
def api_alumno_qr(id):
    a = Alumno.query.get_or_404(id)
    svg_b64 = gen_qr_b64_svg({'id': a.id, 'nombre': a.nombre, 'dni': a.dni})
    return jsonify({'qr': svg_b64, 'mime': 'image/svg+xml', 'nombre': a.nombre, 'dni': a.dni})

@app.route('/api/alumnos/qr/img/<int:id>')
def api_alumno_qr_img(id):
    a = Alumno.query.get_or_404(id)
    svg = gen_qr_svg({'id': a.id, 'nombre': a.nombre, 'dni': a.dni})
    buf = io.BytesIO(svg.encode())
    return send_file(buf, mimetype='image/svg+xml',
                     download_name=f'qr_{a.nombre.replace(" ","_")}.svg',
                     as_attachment=True)

# ══════════════════════════════════════════════════════════════════
# ASISTENCIA
# ══════════════════════════════════════════════════════════════════
@app.route('/api/asistencia/hoy', methods=['GET'])
@login_required
def api_asist_hoy():
    hoy = date.today()
    lista = Asistencia.query.filter_by(fecha=hoy).order_by(Asistencia.hora.desc()).all()
    return jsonify([a.to_dict() for a in lista])

@app.route('/api/asistencia/dni', methods=['POST'])
@login_required
def api_asist_dni():
    ultimos3 = request.json.get('dni','')
    hoy = date.today()
    matches = [a for a in Alumno.query.filter(Alumno.estado!='inactivo').all()
               if str(a.dni).strip().endswith(ultimos3)]
    if not matches:
        return jsonify({'ok': False, 'msg': 'No se encontró alumno con ese DNI'}), 404
    resultados = []
    for alumno in matches:
        ya = Asistencia.query.filter_by(alumno_id=alumno.id, fecha=hoy).first()
        if ya:
            resultados.append({'nombre': alumno.nombre, 'ya': True})
            continue
        ast = Asistencia(alumno_id=alumno.id, fecha=hoy,
                         hora=datetime.now().strftime('%H:%M'), metodo='dni')
        db.session.add(ast)
        resultados.append({'nombre': alumno.nombre, 'ya': False})
    db.session.commit()
    return jsonify({'ok': True, 'resultados': resultados})

@app.route('/api/asistencia/qr', methods=['POST'])
@login_required
def api_asist_qr():
    data = request.json.get('data','')
    hoy = date.today()
    try:
        d = json.loads(data)
        alumno = Alumno.query.get(d.get('id'))
    except:
        alumno = None
    if not alumno:
        return jsonify({'ok': False, 'msg': 'QR no reconocido'}), 404
    ya = Asistencia.query.filter_by(alumno_id=alumno.id, fecha=hoy).first()
    if ya:
        return jsonify({'ok': True, 'ya': True, 'nombre': alumno.nombre})
    ast = Asistencia(alumno_id=alumno.id, fecha=hoy,
                     hora=datetime.now().strftime('%H:%M'), metodo='qr')
    db.session.add(ast); db.session.commit()
    return jsonify({'ok': True, 'ya': False, 'nombre': alumno.nombre, 'alumno': alumno.to_dict()})

# ══════════════════════════════════════════════════════════════════
# MOROSOS + EMAIL
# ══════════════════════════════════════════════════════════════════
@app.route('/api/morosos', methods=['GET'])
@login_required
def api_morosos():
    lista = [a.to_dict() for a in Alumno.query.filter(Alumno.estado!='inactivo').all() if a.es_moroso]
    return jsonify(lista)

@app.route('/api/morosos/email', methods=['POST'])
@login_required
def api_morosos_email():
    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart
    alumno_ids = request.json.get('ids', [])
    host  = get_cfg('smtp_host')
    port  = int(get_cfg('smtp_port','587'))
    user  = get_cfg('smtp_user')
    pasw  = get_cfg('smtp_pass')
    gym   = get_cfg('gym_nombre','Iron Gym')
    if not user or not pasw:
        return jsonify({'ok': False, 'msg': 'Configurá el email SMTP en Ajustes primero'}), 400
    enviados = 0; errores = []
    try:
        server = smtplib.SMTP(host, port)
        server.starttls(); server.login(user, pasw)
        for aid in alumno_ids:
            a = Alumno.query.get(aid)
            if not a or not a.email: continue
            msg = MIMEMultipart('alternative')
            msg['Subject'] = f'[{gym}] Aviso de cuota vencida'
            msg['From'] = user; msg['To'] = a.email
            html = f"""<div style="font-family:Arial,sans-serif;max-width:520px;">
              <h2 style="color:#c0392b;">⚠️ Cuota Vencida — {gym}</h2>
              <p>Hola <strong>{a.nombre}</strong>,</p>
              <p>Te recordamos que tu cuota se encuentra vencida desde hace 
              <strong>{a.dias_vencido} día(s)</strong>.</p>
              <p>Monto adeudado: <strong>${a.monto:,.2f}</strong></p>
              <p>Por favor regularizá tu situación para continuar entrenando.</p>
              <p style="color:#666;">— El equipo de {gym}</p>
            </div>"""
            msg.attach(MIMEText(html,'html'))
            try: server.sendmail(user, a.email, msg.as_string()); enviados += 1
            except Exception as e: errores.append(f'{a.nombre}: {e}')
        server.quit()
    except Exception as e:
        return jsonify({'ok': False, 'msg': str(e)}), 500
    return jsonify({'ok': True, 'enviados': enviados, 'errores': errores})

# ══════════════════════════════════════════════════════════════════
# CLASES
# ══════════════════════════════════════════════════════════════════
@app.route('/api/clases', methods=['GET','POST'])
@login_required
def api_clases():
    if request.method=='GET':
        return jsonify([c.to_dict() for c in Clase.query.filter_by(activa=True).all()])
    d=request.json
    c=Clase(nombre=d['nombre'],instructor=d.get('instructor'),dia=d.get('dia'),
            hora=d.get('hora'),duracion=int(d.get('duracion',60)),
            cupo=int(d.get('cupo',20)),desc=d.get('desc',''))
    db.session.add(c); db.session.commit(); return jsonify(c.to_dict()), 201

@app.route('/api/clases/<int:id>', methods=['PUT','DELETE'])
@login_required
def api_clase(id):
    c=Clase.query.get_or_404(id)
    if request.method=='DELETE':
        c.activa=False; db.session.commit(); return jsonify({'ok':True})
    d=request.json
    c.nombre=d['nombre']; c.instructor=d.get('instructor'); c.dia=d.get('dia')
    c.hora=d.get('hora'); c.duracion=int(d.get('duracion',60))
    c.cupo=int(d.get('cupo',20)); c.desc=d.get('desc','')
    db.session.commit(); return jsonify(c.to_dict())

# ══════════════════════════════════════════════════════════════════
# PRODUCTOS
# ══════════════════════════════════════════════════════════════════
@app.route('/api/productos', methods=['GET','POST'])
@login_required
def api_productos():
    if request.method=='GET':
        cat=request.args.get('cat',''); q=request.args.get('q','').lower()
        query=Producto.query.filter_by(activo=True)
        if cat: query=query.filter_by(categoria=cat)
        if q:   query=query.filter(Producto.nombre.ilike(f'%{q}%'))
        return jsonify([p.to_dict() for p in query.all()])
    d=request.json
    p=Producto(nombre=d['nombre'],categoria=d.get('categoria','otros'),
               emoji=d.get('emoji','📦'),precio=float(d.get('precio',0)),
               stock=int(d.get('stock',0)),stock_min=int(d.get('stock_min',5)),
               desc=d.get('desc',''))
    db.session.add(p); db.session.commit(); return jsonify(p.to_dict()), 201

@app.route('/api/productos/<int:id>', methods=['PUT','DELETE'])
@login_required
def api_producto(id):
    p=Producto.query.get_or_404(id)
    if request.method=='DELETE':
        p.activo=False; db.session.commit(); return jsonify({'ok':True})
    d=request.json
    p.nombre=d['nombre']; p.categoria=d.get('categoria'); p.emoji=d.get('emoji','📦')
    p.precio=float(d.get('precio',0)); p.stock=int(d.get('stock',0))
    p.stock_min=int(d.get('stock_min',5)); p.desc=d.get('desc','')
    db.session.commit(); return jsonify(p.to_dict())

# ══════════════════════════════════════════════════════════════════
# VENTAS
# ══════════════════════════════════════════════════════════════════
@app.route('/api/ventas', methods=['GET','POST'])
@login_required
def api_ventas():
    if request.method=='GET':
        fecha=request.args.get('fecha', date.today().isoformat())
        mes  =request.args.get('mes','')
        if mes:
            y,m=mes.split('-')
            ventas=Venta.query.filter(db.func.strftime('%Y-%m',Venta.fecha)==mes).order_by(Venta.fecha.desc()).all()
        else:
            ventas=Venta.query.filter_by(fecha=date.fromisoformat(fecha)).order_by(Venta.hora.desc()).all()
        return jsonify([v.to_dict() for v in ventas])
    d=request.json; items=d.get('items',[])
    if not items: return jsonify({'error':'Carrito vacío'}), 400
    # descontar stock
    for it in items:
        prod=Producto.query.get(it['id'])
        if prod: prod.stock = max(0, prod.stock - it['qty'])
    v=Venta(fecha=date.today(), hora=datetime.now().strftime('%H:%M'),
            cliente=d.get('cliente','—'), total=float(d.get('total',0)),
            medio_pago=d.get('medio_pago','efectivo'),
            items_json=json.dumps(items), created_by=session.get('user',''))
    db.session.add(v); db.session.commit()
    return jsonify(v.to_dict()), 201

@app.route('/api/ventas/<int:id>', methods=['DELETE'])
@login_required
@admin_required
def api_venta_delete(id):
    v=Venta.query.get_or_404(id); db.session.delete(v); db.session.commit()
    return jsonify({'ok':True})

@app.route('/api/ventas/<int:id>/ticket')
@login_required
def api_ticket(id):
    v=Venta.query.get_or_404(id)
    gym=get_cfg('gym_nombre','Iron Gym'); dir_=get_cfg('gym_dir'); tel=get_cfg('gym_tel')
    html=f"""<!DOCTYPE html><html><head><meta charset="UTF-8">
    <style>body{{font-family:monospace;width:280px;margin:0 auto;font-size:13px;}}
    .center{{text-align:center;}} .line{{border-top:1px dashed #000;margin:6px 0;}}
    .total{{font-size:16px;font-weight:bold;}} </style></head><body>
    <div class="center"><h2 style="margin:4px 0">{gym}</h2>
    <p style="margin:2px 0">{dir_}</p><p style="margin:2px 0">{tel}</p></div>
    <div class="line"></div>
    <p>Fecha: {v.fecha.strftime('%d/%m/%Y')} {v.hora}</p>
    <p>Cliente: {v.cliente}</p><p>Pago: {v.medio_pago}</p>
    <div class="line"></div>
    {''.join(f"<p>{it['nombre']} x{it['qty']} ........... ${it['subtotal']:.2f}</p>" for it in v.items)}
    <div class="line"></div>
    <p class="total center">TOTAL: ${v.total:.2f}</p>
    <div class="line"></div>
    <div class="center"><p>¡Gracias por tu visita!</p></div>
    <script>window.print();</script></body></html>"""
    return html

# ══════════════════════════════════════════════════════════════════
# GASTOS
# ══════════════════════════════════════════════════════════════════
@app.route('/api/gastos', methods=['GET','POST'])
@login_required
@admin_required
def api_gastos():
    if request.method=='GET':
        mes=request.args.get('mes', date.today().strftime('%Y-%m'))
        cat=request.args.get('cat','')
        q=Gasto.query.filter(db.func.strftime('%Y-%m',Gasto.fecha)==mes)
        if cat: q=q.filter_by(categoria=cat)
        return jsonify([g.to_dict() for g in q.order_by(Gasto.fecha.desc()).all()])
    d=request.json
    g=Gasto(fecha=date.fromisoformat(d.get('fecha',date.today().isoformat())),
            desc=d['desc'],categoria=d.get('categoria','otros'),
            monto=float(d.get('monto',0)),comprobante=d.get('comprobante',''),
            created_by=session.get('user',''))
    db.session.add(g); db.session.commit(); return jsonify(g.to_dict()), 201

@app.route('/api/gastos/<int:id>', methods=['PUT','DELETE'])
@login_required
@admin_required
def api_gasto(id):
    g=Gasto.query.get_or_404(id)
    if request.method=='DELETE':
        db.session.delete(g); db.session.commit(); return jsonify({'ok':True})
    d=request.json
    g.fecha=date.fromisoformat(d.get('fecha',g.fecha.isoformat()))
    g.desc=d['desc']; g.categoria=d.get('categoria'); g.monto=float(d.get('monto',0))
    g.comprobante=d.get('comprobante','')
    db.session.commit(); return jsonify(g.to_dict())

# ══════════════════════════════════════════════════════════════════
# PLANES
# ══════════════════════════════════════════════════════════════════
@app.route('/api/planes', methods=['GET','POST'])
@login_required
def api_planes():
    if request.method=='GET':
        return jsonify([{'id':p.id,'nombre':p.nombre,'precio':p.precio}
                        for p in Plan.query.filter_by(activo=True).all()])
    d=request.json
    p=Plan(nombre=d['nombre'],precio=float(d.get('precio',0)))
    db.session.add(p); db.session.commit()
    return jsonify({'id':p.id,'nombre':p.nombre,'precio':p.precio}), 201

@app.route('/api/planes/<int:id>', methods=['PUT','DELETE'])
@login_required
@admin_required
def api_plan(id):
    p=Plan.query.get_or_404(id)
    if request.method=='DELETE':
        p.activo=False; db.session.commit(); return jsonify({'ok':True})
    d=request.json; p.nombre=d['nombre']; p.precio=float(d.get('precio',0))
    db.session.commit(); return jsonify({'id':p.id,'nombre':p.nombre,'precio':p.precio})

# ══════════════════════════════════════════════════════════════════
# CONFIG
# ══════════════════════════════════════════════════════════════════
@app.route('/api/config', methods=['GET'])
@login_required
def api_config_get():
    return jsonify({
        'gym_nombre':  get_cfg('gym_nombre','Iron Gym'),
        'gym_dir':     get_cfg('gym_dir'),
        'gym_tel':     get_cfg('gym_tel'),
        'gym_email':   get_cfg('gym_email'),
        'smtp_host':   get_cfg('smtp_host','smtp.gmail.com'),
        'smtp_port':   get_cfg('smtp_port','587'),
        'smtp_user':   get_cfg('smtp_user'),
        'fondo_tipo':  get_cfg('fondo_tipo','color'),
        'fondo_valor': get_cfg('fondo_valor','#0a0a0a'),
    })

@app.route('/api/config', methods=['POST'])
@login_required
@admin_required
def api_config_save():
    d = request.json
    for k in ['gym_nombre','gym_dir','gym_tel','gym_email','smtp_host','smtp_port','smtp_user','fondo_tipo','fondo_valor']:
        if k in d: set_cfg(k, d[k])
    if d.get('smtp_pass'): set_cfg('smtp_pass', d['smtp_pass'])
    if d.get('pass_admin'): set_cfg('pass_admin', hash_pass(d['pass_admin']))
    if d.get('pass_op'):    set_cfg('pass_op',    hash_pass(d['pass_op']))
    if d.get('master_key'): set_cfg('master_key', hash_pass(d['master_key']))
    return jsonify({'ok': True})

# ── Fondo imagen upload ──────────────────────────────────────────
@app.route('/api/config/fondo', methods=['POST'])
@login_required
@admin_required
def api_fondo_upload():
    if 'imagen' not in request.files:
        return jsonify({'ok': False, 'msg': 'No se recibió archivo'}), 400
    f = request.files['imagen']
    if not f.filename:
        return jsonify({'ok': False, 'msg': 'Archivo vacío'}), 400
    allowed = {'png','jpg','jpeg','gif','webp'}
    ext = f.filename.rsplit('.',1)[-1].lower()
    if ext not in allowed:
        return jsonify({'ok': False, 'msg': 'Formato no permitido. Usar PNG, JPG o WEBP'}), 400
    data = f.read()
    if len(data) > 5 * 1024 * 1024:
        return jsonify({'ok': False, 'msg': 'Imagen demasiado grande (máx 5 MB)'}), 400
    b64 = base64.b64encode(data).decode()
    mime = f'image/{ext}' if ext != 'jpg' else 'image/jpeg'
    data_url = f'data:{mime};base64,{b64}'
    set_cfg('fondo_tipo',  'imagen')
    set_cfg('fondo_valor', data_url)
    return jsonify({'ok': True, 'fondo_valor': data_url})

# ══════════════════════════════════════════════════════════════════
# IMPORTACIÓN MASIVA — EXCEL
# ══════════════════════════════════════════════════════════════════
@app.route('/api/import/alumnos', methods=['POST'])
@login_required
@admin_required
def api_import_alumnos():
    import openpyxl
    if 'archivo' not in request.files:
        return jsonify({'ok': False, 'msg': 'No se recibió archivo'}), 400
    f = request.files['archivo']
    ext = f.filename.rsplit('.',1)[-1].lower()
    creados = 0; omitidos = 0; errores = []

    if ext == 'xlsx':
        wb = openpyxl.load_workbook(f, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows: return jsonify({'ok':False,'msg':'Archivo vacío'}), 400
        # Detect header row — skip if first cell is text like 'nombre'
        start = 1 if str(rows[0][0]).lower().strip() in ('nombre','name','alumno') else 0
        for row in rows[start:]:
            if not row or not row[0]: continue
            try:
                nombre  = str(row[0]).strip()
                dni     = str(int(row[1])) if row[1] is not None else ''
                tel     = str(row[2]).strip() if len(row)>2 and row[2] else ''
                email   = str(row[3]).strip() if len(row)>3 and row[3] else ''
                plan_nm = str(row[4]).strip() if len(row)>4 and row[4] else ''
                monto   = float(row[5]) if len(row)>5 and row[5] else 0
                vto_raw = row[6] if len(row)>6 else None
                vto     = None
                if vto_raw:
                    if hasattr(vto_raw,'date'): vto = vto_raw.date()
                    else:
                        try:
                            from datetime import datetime as dt_
                            vto = dt_.strptime(str(vto_raw),'%d/%m/%Y').date()
                        except: pass
                if not nombre or not dni:
                    omitidos += 1; continue
                if Alumno.query.filter_by(dni=dni).first():
                    omitidos += 1; continue
                plan = Plan.query.filter(Plan.nombre.ilike(f'%{plan_nm}%')).first() if plan_nm else None
                a = Alumno(nombre=nombre, dni=dni, tel=tel, email=email,
                           plan_id=plan.id if plan else None,
                           monto=monto, vto_cuota=vto,
                           alta=date.today(), estado='activo')
                db.session.add(a); creados += 1
            except Exception as e:
                errores.append(f'Fila {rows.index(row)+1}: {str(e)}')
    elif ext == 'csv':
        import csv as csv_mod, io
        content = f.read().decode('utf-8-sig')
        reader = csv_mod.DictReader(io.StringIO(content))
        for row in reader:
            try:
                nombre = (row.get('nombre') or row.get('Nombre') or '').strip()
                dni    = (row.get('dni')    or row.get('DNI')    or '').strip()
                if not nombre or not dni: omitidos+=1; continue
                if Alumno.query.filter_by(dni=dni).first(): omitidos+=1; continue
                a = Alumno(nombre=nombre,dni=dni,
                           tel=(row.get('tel') or row.get('telefono') or '').strip(),
                           email=(row.get('email') or '').strip(),
                           monto=float(row.get('monto') or 0),
                           alta=date.today(), estado='activo')
                db.session.add(a); creados += 1
            except Exception as e:
                errores.append(str(e))
    else:
        return jsonify({'ok':False,'msg':'Formato no soportado. Usar .xlsx o .csv'}), 400

    db.session.commit()
    return jsonify({'ok':True,'creados':creados,'omitidos':omitidos,'errores':errores})

@app.route('/api/import/productos', methods=['POST'])
@login_required
@admin_required
def api_import_productos():
    import openpyxl
    if 'archivo' not in request.files:
        return jsonify({'ok': False, 'msg': 'No se recibió archivo'}), 400
    f = request.files['archivo']
    ext = f.filename.rsplit('.',1)[-1].lower()
    creados = 0; omitidos = 0; errores = []
    cats_validas = {'equipamiento','bebidas','rehidratantes','cuotas','otros'}

    if ext == 'xlsx':
        wb = openpyxl.load_workbook(f, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows: return jsonify({'ok':False,'msg':'Archivo vacío'}), 400
        start = 1 if str(rows[0][0]).lower().strip() in ('nombre','name','producto') else 0
        for row in rows[start:]:
            if not row or not row[0]: continue
            try:
                nombre  = str(row[0]).strip()
                cat     = str(row[1]).strip().lower() if len(row)>1 and row[1] else 'otros'
                precio  = float(row[2]) if len(row)>2 and row[2] else 0
                stock   = int(row[3])   if len(row)>3 and row[3] else 0
                stockm  = int(row[4])   if len(row)>4 and row[4] else 5
                emoji   = str(row[5]).strip() if len(row)>5 and row[5] else '📦'
                if cat not in cats_validas: cat = 'otros'
                if not nombre: omitidos+=1; continue
                p = Producto(nombre=nombre, categoria=cat, precio=precio,
                             stock=stock, stock_min=stockm, emoji=emoji)
                db.session.add(p); creados += 1
            except Exception as e:
                errores.append(str(e))
    elif ext == 'csv':
        import csv as csv_mod, io
        content = f.read().decode('utf-8-sig')
        reader = csv_mod.DictReader(io.StringIO(content))
        for row in reader:
            try:
                nombre = (row.get('nombre') or row.get('Nombre') or '').strip()
                if not nombre: omitidos+=1; continue
                cat = (row.get('categoria') or row.get('Categoria') or 'otros').strip().lower()
                if cat not in cats_validas: cat = 'otros'
                p = Producto(nombre=nombre, categoria=cat,
                             precio=float(row.get('precio') or 0),
                             stock=int(row.get('stock') or 0),
                             stock_min=int(row.get('stock_min') or 5),
                             emoji=row.get('emoji','📦'))
                db.session.add(p); creados += 1
            except Exception as e:
                errores.append(str(e))
    else:
        return jsonify({'ok':False,'msg':'Formato no soportado. Usar .xlsx o .csv'}), 400

    db.session.commit()
    return jsonify({'ok':True,'creados':creados,'omitidos':omitidos,'errores':errores})

# ── Template download ────────────────────────────────────────────
@app.route('/api/import/template/<tipo>')
@login_required
def api_import_template(tipo):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    hdr_font  = Font(bold=True, color='FFFFFF', name='Arial', size=11)
    hdr_fill  = PatternFill('solid', fgColor='C0392B')
    ex_fill   = PatternFill('solid', fgColor='F5F5F5')
    center    = Alignment(horizontal='center', vertical='center')

    if tipo == 'alumnos':
        ws.title = 'Alumnos'
        headers = ['nombre','dni','telefono','email','plan','monto_cuota','vto_cuota (dd/mm/aaaa)']
        ejemplos = [
            ['Juan Pérez','12345678','1145678901','juan@mail.com','3 días/semana',8000,'31/12/2025'],
            ['María García','23456789','1156789012','','Libre',15000,''],
        ]
        col_w = [28,14,18,28,20,16,24]
    else:
        ws.title = 'Productos'
        headers = ['nombre','categoria','precio','stock','stock_minimo','emoji']
        ejemplos = [
            ['Guantes 12oz','equipamiento',8500,15,3,'🥊'],
            ['Agua 500ml','bebidas',350,50,10,'💧'],
            ['Isotónico','rehidratantes',800,20,5,'⚡'],
            ['Cuota Mensual','cuotas',12000,0,0,'💳'],
        ]
        col_w = [28,18,12,10,14,8]
        # Add category note
        ws['H2'] = 'Categorías válidas:'
        ws['H3'] = 'equipamiento'
        ws['H4'] = 'bebidas'
        ws['H5'] = 'rehidratantes'
        ws['H6'] = 'cuotas'
        ws['H7'] = 'otros'

    for ci, (h, w) in enumerate(zip(headers, col_w), 1):
        cell = ws.cell(1, ci, h)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = center
        ws.column_dimensions[cell.column_letter].width = w

    for ri, row in enumerate(ejemplos, 2):
        for ci, v in enumerate(row, 1):
            cell = ws.cell(ri, ci, v)
            cell.fill = ex_fill
            cell.alignment = Alignment(vertical='center')

    ws.row_dimensions[1].height = 22

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    fname = f'template_{tipo}.xlsx'
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     download_name=fname, as_attachment=True)

@app.route('/api/config/test-email', methods=['POST'])
@login_required
@admin_required
def api_test_email():
    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart
    host  = get_cfg('smtp_host', 'smtp.gmail.com')
    port  = int(get_cfg('smtp_port', '587'))
    user  = get_cfg('smtp_user', '')
    pasw  = get_cfg('smtp_pass', '')
    gym   = get_cfg('gym_nombre', 'Iron Gym')
    dest  = request.json.get('destino', '').strip()
    if not user or not pasw:
        return jsonify({'ok': False, 'msg': 'Completá el email y contraseña SMTP antes de probar.'}), 400
    if not dest:
        return jsonify({'ok': False, 'msg': 'Ingresá un email de destino.'}), 400
    try:
        server = smtplib.SMTP(host, port, timeout=10)
        server.ehlo()
        server.starttls()
        server.ehlo()
        server.login(user, pasw)
        msg = MIMEMultipart('alternative')
        msg['Subject'] = f'[{gym}] ✅ Email de prueba — configuración correcta'
        msg['From']    = user
        msg['To']      = dest
        html = f"""
        <div style="font-family:Arial,sans-serif;max-width:480px;padding:20px;border:1px solid #ddd">
          <h2 style="color:#c0392b">🥊 {gym}</h2>
          <p>Este es un <strong>email de prueba</strong> enviado desde el sistema de gestión.</p>
          <p style="color:#27ae60;font-weight:bold">✅ La configuración de email funciona correctamente.</p>
          <hr style="border:none;border-top:1px solid #eee;margin:16px 0">
          <p style="color:#999;font-size:12px">Iron Gym — Sistema de Gestión</p>
        </div>"""
        msg.attach(MIMEText(html, 'html'))
        server.sendmail(user, dest, msg.as_string())
        server.quit()
        return jsonify({'ok': True, 'msg': f'Email enviado correctamente a {dest}'})
    except smtplib.SMTPAuthenticationError:
        return jsonify({'ok': False, 'msg': 'Error de autenticación. Verificá usuario y contraseña (o App Password para Gmail).'}), 400
    except smtplib.SMTPConnectError:
        return jsonify({'ok': False, 'msg': f'No se pudo conectar a {host}:{port}. Verificá el servidor y puerto.'}), 400
    except TimeoutError:
        return jsonify({'ok': False, 'msg': f'Timeout al conectar con {host}:{port}. Verificá tu conexión a internet.'}), 400
    except Exception as e:
        return jsonify({'ok': False, 'msg': f'Error: {str(e)}'}), 500

# ══════════════════════════════════════════════════════════════════
# EXPORTACIONES CSV
# ══════════════════════════════════════════════════════════════════
@app.route('/api/export/ventas')
@login_required
def exp_ventas():
    mes=request.args.get('mes', date.today().strftime('%Y-%m'))
    ventas=Venta.query.filter(db.func.strftime('%Y-%m',Venta.fecha)==mes).all()
    buf=io.StringIO()
    w=csv.writer(buf)
    w.writerow(['ID','Fecha','Hora','Items','Cliente','Total','Pago'])
    for v in ventas:
        items_str=' | '.join(f"{i['nombre']} x{i['qty']}" for i in v.items)
        w.writerow([v.id, v.fecha, v.hora, items_str, v.cliente, v.total, v.medio_pago])
    buf.seek(0)
    return send_file(io.BytesIO(buf.getvalue().encode()),
                     mimetype='text/csv', download_name=f'ventas_{mes}.csv', as_attachment=True)

@app.route('/api/export/alumnos')
@login_required
def exp_alumnos():
    buf=io.StringIO(); w=csv.writer(buf)
    w.writerow(['ID','Nombre','DNI','Teléfono','Email','Plan','Estado','Vencimiento','Monto'])
    for a in Alumno.query.order_by(Alumno.nombre).all():
        w.writerow([a.id,a.nombre,a.dni,a.tel or '',a.email or '',
                    a.plan.nombre if a.plan else '',a.estado,
                    a.vto_cuota or '',a.monto])
    buf.seek(0)
    return send_file(io.BytesIO(buf.getvalue().encode()),
                     mimetype='text/csv', download_name='alumnos.csv', as_attachment=True)

@app.route('/api/export/asistencias')
@login_required
def exp_asistencias():
    mes=request.args.get('mes', date.today().strftime('%Y-%m'))
    lista=Asistencia.query.filter(db.func.strftime('%Y-%m',Asistencia.fecha)==mes).all()
    buf=io.StringIO(); w=csv.writer(buf)
    w.writerow(['ID','Fecha','Hora','Alumno','DNI','Método'])
    for a in lista:
        w.writerow([a.id,a.fecha,a.hora,
                    a.alumno.nombre if a.alumno else '',
                    a.alumno.dni if a.alumno else '',a.metodo])
    buf.seek(0)
    return send_file(io.BytesIO(buf.getvalue().encode()),
                     mimetype='text/csv', download_name=f'asistencias_{mes}.csv', as_attachment=True)

@app.route('/api/export/gastos')
@login_required
@admin_required
def exp_gastos():
    mes=request.args.get('mes', date.today().strftime('%Y-%m'))
    lista=Gasto.query.filter(db.func.strftime('%Y-%m',Gasto.fecha)==mes).all()
    buf=io.StringIO(); w=csv.writer(buf)
    w.writerow(['ID','Fecha','Descripción','Categoría','Monto','Comprobante'])
    for g in lista:
        w.writerow([g.id,g.fecha,g.desc,g.categoria,g.monto,g.comprobante or ''])
    buf.seek(0)
    return send_file(io.BytesIO(buf.getvalue().encode()),
                     mimetype='text/csv', download_name=f'gastos_{mes}.csv', as_attachment=True)

# ══════════════════════════════════════════════════════════════════
# REPORTES
# ══════════════════════════════════════════════════════════════════
@app.route('/api/reportes/resumen')
@login_required
@admin_required
def api_reporte_resumen():
    mes=request.args.get('mes', date.today().strftime('%Y-%m'))
    ventas=Venta.query.filter(db.func.strftime('%Y-%m',Venta.fecha)==mes).all()
    gastos=Gasto.query.filter(db.func.strftime('%Y-%m',Gasto.fecha)==mes).all()
    total_v=sum(v.total for v in ventas)
    total_g=sum(g.monto for g in gastos)
    cats={}
    for v in ventas:
        for it in v.items:
            c=it.get('cat','otros')
            if c not in cats: cats[c]={'qty':0,'total':0}
            cats[c]['qty']+=it.get('qty',1); cats[c]['total']+=it.get('subtotal',0)
    return jsonify({'total_ventas':total_v,'total_gastos':total_g,'neto':total_v-total_g,
                    'tickets':len(ventas),'cats':cats})

# ══════════════════════════════════════════════════════════════════
# ENTRY POINT
# ══════════════════════════════════════════════════════════════════
if __name__ == '__main__':
    import socket
    with app.app_context():
        db.create_all()
        seed_defaults()
        run_migrations()
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(('8.8.8.8', 80)); local_ip = s.getsockname()[0]; s.close()
    except: local_ip = '127.0.0.1'

    env_label = "TEST" if IS_TEST else "PRODUCCION"
    print("\n" + "="*52)
    print(f"  IRON GYM v{APP_VERSION}  [{env_label}]")
    print("="*52)
    print(f"  Local:    http://127.0.0.1:{APP_PORT}")
    print(f"  Red LAN:  http://{local_ip}:{APP_PORT}")
    if IS_TEST:
        print(f"  ENTORNO TEST — DB separada")
        print(f"  instance_test/irongym_test.db")
    print(f"\n  admin:    admin / admin123")
    print(f"  operador: operador / op123")
    print(f"  master:   IRONGYM2024")
    print("="*52 + "\n")

    if IS_TEST:
        # Test: servidor de desarrollo Flask (debug activo)
        app.run(host='0.0.0.0', port=APP_PORT, debug=True)
    else:
        # Producción: Waitress — servidor WSGI estable para Windows
        try:
            from waitress import serve
            print(f"  Servidor: Waitress (produccion)")
            print(f"  Threads:  8")
            print("="*52 + "\n")
            serve(app, host='0.0.0.0', port=APP_PORT, threads=8,
                  channel_timeout=60, cleanup_interval=30)
        except ImportError:
            # Waitress no instalado — fallback a Flask dev server con advertencia
            print("  ADVERTENCIA: waitress no instalado, usando servidor de desarrollo.")
            print("  Instalar con: pip install waitress")
            print("="*52 + "\n")
            app.run(host='0.0.0.0', port=APP_PORT, debug=False)
